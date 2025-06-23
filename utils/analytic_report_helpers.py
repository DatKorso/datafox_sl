"""
Helper functions for analytic report processing.

This module provides functionality to:
- Load and validate analytic report Excel files
- Map WB SKUs to Ozon SKUs by sizes
- Calculate size ranges and stock aggregations
- Generate order statistics for analytic reports
- Update Excel files with calculated data while preserving structure
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
from datetime import datetime, timedelta
import streamlit as st
from typing import Dict, List, Tuple, Optional
import requests
from io import BytesIO
import tempfile
from PIL import Image as PILImage
from utils.db_search_helpers import get_normalized_wb_barcodes, get_ozon_barcodes_and_identifiers

def load_analytic_report_file(file_path: str) -> Tuple[Optional[pd.DataFrame], Optional[openpyxl.Workbook], str]:
    """
    Loads and validates an analytic report Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Tuple of (DataFrame with data from row 9+, Workbook object, error message if any)
    """
    if not os.path.exists(file_path):
        return None, None, f"Файл не найден: {file_path}"
    
    try:
        # Load workbook to check structure
        wb = load_workbook(file_path)
        
        if "analytic_report" not in wb.sheetnames:
            return None, None, "Лист 'analytic_report' не найден в файле"
        
        ws = wb["analytic_report"]
        
        # Check if there are at least 9 rows
        if ws.max_row < 9:
            return None, None, "В файле недостаточно строк. Ожидается минимум 9 строк"
        
        # Load data starting from row 9 (index 8 in pandas)
        df = pd.read_excel(file_path, sheet_name="analytic_report", header=6, skiprows=[7])
        
        # Check if WB_SKU column exists
        if "WB_SKU" not in df.columns:
            return None, None, "Колонка 'WB_SKU' не найдена в 7-й строке файла"
        
        # Remove empty rows
        df = df.dropna(subset=["WB_SKU"])
        
        if df.empty:
            return None, None, "Нет данных с WB_SKU для обработки"
        
        return df, wb, ""
        
    except Exception as e:
        return None, None, f"Ошибка при загрузке файла: {str(e)}"

def map_wb_to_ozon_by_size(db_conn, wb_sku_list: List[str]) -> Dict[str, Dict[int, List[str]]]:
    """
    Maps WB SKUs to Ozon SKUs grouped by sizes.
    Uses wb_products.wb_size to determine size and maps through common barcodes.
    
    ОБНОВЛЕНО: Теперь использует централизованный CrossMarketplaceLinker
    для исключения дублирования логики связывания.
    
    Args:
        db_conn: Database connection
        wb_sku_list: List of WB SKUs to process
        
    Returns:
        Dict mapping wb_sku -> size -> [list of oz_skus]
    """
    if not wb_sku_list:
        return {}
    
    try:
        # Use centralized linker for WB-Ozon links grouped by sizes
        from utils.cross_marketplace_linker import CrossMarketplaceLinker
        linker = CrossMarketplaceLinker(db_conn)
        
        return linker.get_links_with_sizes(wb_sku_list)
        
    except Exception as e:
        st.error(f"Ошибка при сопоставлении WB и Ozon размеров: {e}")
        return {}

def calculate_size_range(size_mapping: Dict[int, List[str]]) -> str:
    """
    Calculates size range string from size mapping.
    
    Args:
        size_mapping: Dict mapping size -> [list of oz_skus]
        
    Returns:
        Size range string like "27-38" or empty string if no sizes
    """
    sizes_with_skus = [size for size, skus in size_mapping.items() if skus]
    
    if not sizes_with_skus:
        return ""
    
    min_size = min(sizes_with_skus)
    max_size = max(sizes_with_skus)
    
    if min_size == max_size:
        return str(min_size)
    else:
        return f"{min_size}-{max_size}"

def aggregate_stock_data(db_conn, oz_sku_list: List[str]) -> int:
    """
    Aggregates stock data for a list of Ozon SKUs.
    
    Args:
        db_conn: Database connection
        oz_sku_list: List of Ozon SKUs
        
    Returns:
        Total stock amount
    """
    if not oz_sku_list:
        return 0
    
    try:
        stock_query = f"""
        SELECT COALESCE(SUM(oz_fbo_stock), 0) as total_stock
        FROM oz_products 
        WHERE oz_sku IN ({', '.join(['?'] * len(oz_sku_list))})
        """
        result = db_conn.execute(stock_query, oz_sku_list).fetchdf()
        return int(result.iloc[0]['total_stock']) if not result.empty else 0
    except Exception as e:
        st.error(f"Ошибка при получении данных о остатках: {e}")
        return 0

def generate_order_statistics(db_conn, oz_sku_list: List[str], days_back: int = 30) -> Dict[str, int]:
    """
    Generates order statistics for Ozon SKUs for the last N days.
    
    Args:
        db_conn: Database connection
        oz_sku_list: List of Ozon SKUs
        days_back: Number of days to look back
        
    Returns:
        Dict mapping date_string (YYYY-MM-DD) -> order_count
    """
    if not oz_sku_list:
        return {}
    
    try:
        today = datetime.now()
        start_date = (today - timedelta(days=days_back)).strftime('%Y-%m-%d')
        
        orders_query = f"""
        SELECT 
            oz_accepted_date,
            COUNT(*) as order_count
        FROM oz_orders
        WHERE oz_sku IN ({', '.join(['?'] * len(oz_sku_list))})
            AND oz_accepted_date >= ?
            AND order_status != 'Отменён'
        GROUP BY oz_accepted_date
        ORDER BY oz_accepted_date DESC
        """
        
        orders_df = db_conn.execute(orders_query, oz_sku_list + [start_date]).fetchdf()
        
        if orders_df.empty:
            return {}
        
        # Convert to dict
        orders_df['oz_accepted_date'] = pd.to_datetime(orders_df['oz_accepted_date']).dt.strftime('%Y-%m-%d')
        return dict(zip(orders_df['oz_accepted_date'], orders_df['order_count']))
        
    except Exception as e:
        st.error(f"Ошибка при получении статистики заказов: {e}")
        return {}

def create_backup_file(file_path: str) -> str:
    """
    Creates a backup of the original file with timestamp.
    
    Args:
        file_path: Path to the original file
        
    Returns:
        Path to the backup file
    """
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    name, ext = os.path.splitext(filename)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"{name}_backup_{timestamp}{ext}"
    backup_path = os.path.join(directory, backup_filename)
    
    # Copy the file
    import shutil
    shutil.copy2(file_path, backup_path)
    
    return backup_path

def get_wb_image_url(wb_sku: str) -> str:
    """
    Генерирует URL изображения WB на основе артикула.
    Реализует алгоритм из Google Sheets для определения правильного URL.
    
    Args:
        wb_sku: Артикул WB (строка или число)
        
    Returns:
        URL изображения или пустая строка если не удалось сгенерировать
    """
    try:
        # Конвертируем в целое число
        a = int(float(wb_sku))
        
        # Вычисляем b и c
        b = a // 100000
        c = a // 1000
        
        # Определяем volNum по алгоритму из Google Sheets
        if b <= 143:
            vol_num = "01"
        elif b <= 287:
            vol_num = "02"
        elif b <= 431:
            vol_num = "03"
        elif b <= 719:
            vol_num = "04"
        elif b <= 1007:
            vol_num = "05"
        elif b <= 1061:
            vol_num = "06"
        elif b <= 1115:
            vol_num = "07"
        elif b <= 1169:
            vol_num = "08"
        elif b <= 1313:
            vol_num = "09"
        elif b <= 1601:
            vol_num = "10"
        elif b <= 1655:
            vol_num = "11"
        elif b <= 1919:
            vol_num = "12"
        elif b <= 2045:
            vol_num = "13"
        elif b <= 2189:
            vol_num = "14"
        elif b <= 2405:
            vol_num = "15"
        elif b <= 2621:
            vol_num = "16"
        elif b <= 2837:
            vol_num = "17"
        elif b <= 3053:
            vol_num = "18"
        elif b <= 3269:
            vol_num = "19"
        elif b <= 3485:
            vol_num = "20"
        elif b <= 3701:
            vol_num = "21"
        elif b <= 3917:
            vol_num = "22"
        elif b <= 4133:
            vol_num = "23"
        elif b <= 4349:
            vol_num = "24"
        else:
            vol_num = "25"
        
        # Формируем URL
        image_url = f"https://basket-{vol_num}.wbbasket.ru/vol{b}/part{c}/{a}/images/tm/1.webp"
        return image_url
        
    except (ValueError, TypeError) as e:
        st.warning(f"Не удалось сгенерировать URL изображения для WB_SKU {wb_sku}: {e}")
        return ""

def download_wb_image(wb_sku: str, timeout: int = 30) -> Optional[BytesIO]:
    """
    Загружает изображение WB по артикулу.
    
    Args:
        wb_sku: Артикул WB
        timeout: Таймаут для запроса в секундах
        
    Returns:
        BytesIO объект с изображением или None если загрузка не удалась
    """
    image_url = get_wb_image_url(wb_sku)
    if not image_url:
        return None
    
    try:
        # Используем headers чтобы имитировать браузер
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        
        response = requests.get(image_url, timeout=timeout, headers=headers)
        response.raise_for_status()
        
        # Проверяем что это действительно изображение
        content_type = response.headers.get('content-type', '')
        if not content_type.startswith('image/'):
            st.warning(f"URL не содержит изображение для WB_SKU {wb_sku}: {content_type}")
            return None
        
        return BytesIO(response.content)
        
    except requests.exceptions.RequestException as e:
        st.warning(f"Ошибка при загрузке изображения для WB_SKU {wb_sku}: {e}")
        return None
    except Exception as e:
        st.warning(f"Неожиданная ошибка при загрузке изображения для WB_SKU {wb_sku}: {e}")
        return None

def insert_wb_image_to_cell(ws, row_num: int, col_num: int, wb_sku: str, cell_width: float = 64, cell_height: float = 64) -> Tuple[bool, Optional[str]]:
    """
    Загружает и вставляет изображение WB в ячейку Excel.
    Конвертирует WebP в PNG для совместимости с Excel.
    Привязывает изображение к ячейке для правильного поведения при изменении размеров.
    
    Args:
        ws: Worksheet объект openpyxl
        row_num: Номер строки
        col_num: Номер колонки
        wb_sku: Артикул WB
        cell_width: Ширина изображения в пикселях
        cell_height: Высота изображения в пикселях
        
    Returns:
        Tuple: (успех, путь_к_временному_файлу_для_очистки)
        Возвращает путь к временному файлу, который нужно удалить позже
    """
    try:
        # Загружаем изображение
        image_data = download_wb_image(wb_sku)
        if not image_data:
            return False, None
        
        # Создаем временный файл для изображения в формате PNG
        try:
            temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
            temp_file_path = temp_file.name
            temp_file.close()  # Закрываем файл для записи через PIL
            
            # Открываем WebP изображение из BytesIO
            image_data.seek(0)
            pil_image = PILImage.open(image_data)
            
            # Конвертируем в RGB если необходимо (для PNG)
            if pil_image.mode in ("RGBA", "P"):
                pil_image = pil_image.convert("RGB")
            
            # Сохраняем как PNG
            pil_image.save(temp_file_path, "PNG")
            
            # Дополнительная проверка что файл существует
            if not os.path.exists(temp_file_path):
                st.error(f"❌ Не удалось создать временный файл для WB_SKU {wb_sku}")
                return False, None
            
        except Exception as temp_error:
            st.error(f"❌ Ошибка конвертации изображения для WB_SKU {wb_sku}: {temp_error}")
            return False, None
        
        try:
            # Создаем объект изображения openpyxl
            img = Image(temp_file_path)
            
            # Устанавливаем размер изображения
            img.width = cell_width
            img.height = cell_height
            
            # Определяем координаты ячейки
            from openpyxl.utils import get_column_letter
            cell_address = f"{get_column_letter(col_num)}{row_num}"
            
            # Настраиваем привязку изображения к ячейке
            # Это заставляет изображение двигаться и изменяться вместе с ячейкой
            img.anchor = cell_address
            
            # Вставляем изображение в ячейку (без текста)
            ws.add_image(img, cell_address)
            
            # Настраиваем размер строки и столбца чтобы изображение помещалось
            ws.row_dimensions[row_num].height = max(ws.row_dimensions[row_num].height or 15, cell_height * 0.75)
            
            # Устанавливаем ширину столбца если необходимо
            column_letter = get_column_letter(col_num)
            current_width = ws.column_dimensions[column_letter].width
            if current_width is None or current_width < (cell_width / 7):  # Примерное соотношение пикселей к Excel единицам
                ws.column_dimensions[column_letter].width = cell_width / 7
            
            # Оставляем ячейку пустой (без текста гиперссылки)
            # Если нужна гиперссылка, её можно добавить через комментарий к изображению
            
            return True, temp_file_path
            
        except Exception as e_insert:
            st.error(f"❌ Ошибка вставки изображения для WB_SKU {wb_sku}: {e_insert}")
            # Если вставка не удалась, удаляем временный файл сразу
            try:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
            except Exception as cleanup_error:
                pass  # Тихо игнорируем ошибки очистки
            raise e_insert
                
    except Exception as e:
        st.error(f"❌ Общая ошибка при обработке изображения для WB_SKU {wb_sku}: {e}")
        return False, None

def update_analytic_report(file_path: str, wb_sku_data: Dict[str, Dict], include_images: bool = False) -> Tuple[bool, str]:
    """
    Updates the analytic report Excel file with calculated data.
    Now includes support for PHOTO_FROM_WB column with image insertion.
    
    Args:
        file_path: Path to the Excel file
        wb_sku_data: Dict containing calculated data for each WB SKU
        include_images: Whether to download and insert product images from WB
        
    Returns:
        Tuple of (success, error_message)
    """
    temp_image_files = []  # Список временных файлов для очистки
    
    try:
        # Create backup only for non-temporary files
        backup_created = False
        if not os.path.basename(file_path).startswith('temp_'):
            backup_path = create_backup_file(file_path)
            st.info(f"Создана резервная копия: {os.path.basename(backup_path)}")
            backup_created = True
        
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb["analytic_report"]
        
        # Find column indices from row 7 (header row)
        header_row = 7
        column_map = {}
        
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_num).value
            if cell_value:
                column_map[str(cell_value).strip()] = col_num
        
        # Check if PHOTO_FROM_WB column exists and images are enabled
        has_photo_column = "PHOTO_FROM_WB" in column_map and include_images
        if "PHOTO_FROM_WB" in column_map:
            if include_images:
                st.info("🖼️ Найдена колонка PHOTO_FROM_WB - будут загружены изображения товаров")
            else:
                st.info("📷 Найдена колонка PHOTO_FROM_WB - обновление изображений отключено (ячейки останутся пустыми)")
        
        # Process data starting from row 9
        data_start_row = 9
        photo_success_count = 0
        photo_total_count = 0
        failed_images = []  # Список неудачных загрузок для детального лога
        
        # Collect all WB SKUs for batch processing info
        wb_skus_to_process = []
        for row_num in range(data_start_row, ws.max_row + 1):
            wb_sku_cell = ws.cell(row=row_num, column=column_map.get("WB_SKU", 2))
            wb_sku = str(wb_sku_cell.value).strip() if wb_sku_cell.value else ""
            if wb_sku and wb_sku in wb_sku_data:
                wb_skus_to_process.append((row_num, wb_sku))
        
        # Show compact processing info
        if has_photo_column and wb_skus_to_process:
            st.info(f"🔄 Обработка {len(wb_skus_to_process)} изображений товаров...")
        
        # Process each row
        for row_num, wb_sku in wb_skus_to_process:
            data = wb_sku_data[wb_sku]
            
            # Update size columns (OZ_SIZE_22 to OZ_SIZE_44)
            for size in range(22, 45):
                col_name = f"OZ_SIZE_{size}"
                if col_name in column_map:
                    col_num = column_map[col_name]
                    oz_skus = data.get('sizes', {}).get(size, [])
                    ws.cell(row=row_num, column=col_num).value = ';'.join(oz_skus) if oz_skus else ""
            
            # Update OZ_SIZES
            if "OZ_SIZES" in column_map:
                ws.cell(row=row_num, column=column_map["OZ_SIZES"]).value = data.get('size_range', '')
            
            # Update OZ_STOCK
            if "OZ_STOCK" in column_map:
                ws.cell(row=row_num, column=column_map["OZ_STOCK"]).value = data.get('total_stock', 0)
            
            # Update order statistics columns (ORDERS_TODAY-30 to ORDERS_TODAY-1)
            orders_data = data.get('orders', {})
            today = datetime.now()
            for days_back in range(1, 31):
                col_name = f"ORDERS_TODAY-{days_back}"
                if col_name in column_map:
                    target_date = (today - timedelta(days=days_back)).strftime('%Y-%m-%d')
                    order_count = orders_data.get(target_date, 0)
                    ws.cell(row=row_num, column=column_map[col_name]).value = order_count
            
            # Insert WB image if PHOTO_FROM_WB column exists
            if has_photo_column:
                photo_total_count += 1
                col_num = column_map["PHOTO_FROM_WB"]
                
                # Генерируем URL для проверки доступности изображения
                image_url = get_wb_image_url(wb_sku)
                
                if image_url:
                    # Не добавляем никакого текста в ячейку - только изображение
                    
                    # Пытаемся вставить изображение
                    try:
                        success, temp_file_path = insert_wb_image_to_cell(ws, row_num, col_num, wb_sku)
                        if success:
                            photo_success_count += 1
                            if temp_file_path:
                                temp_image_files.append(temp_file_path)
                        else:
                            failed_images.append(f"WB_SKU {wb_sku} (строка {row_num}): Не удалось загрузить изображение")
                    except Exception as image_error:
                        failed_images.append(f"WB_SKU {wb_sku} (строка {row_num}): {str(image_error)}")
                else:
                    failed_images.append(f"WB_SKU {wb_sku} (строка {row_num}): Не удалось сгенерировать URL")
        
        # Update all PUNTA_ columns dynamically
        punta_columns = {col_name: col_num for col_name, col_num in column_map.items() 
                         if col_name.startswith('PUNTA_')}
        if punta_columns:
            st.info(f"🔄 Обновление {len(punta_columns)} PUNTA_ колонок: {', '.join(punta_columns.keys())}")
        
        update_punta_columns_dynamically(ws, column_map, wb_sku_data, data_start_row)
        
        # Save the updated file
        wb.save(file_path)
        
        # Show compact image loading statistics
        if has_photo_column and photo_total_count > 0:
            if photo_success_count == photo_total_count:
                st.success(f"🖼️ Все изображения загружены успешно: {photo_success_count} из {photo_total_count}")
            else:
                st.warning(f"🖼️ Изображения: {photo_success_count} из {photo_total_count} успешно загружены")
                
                # Show failed images in expander
                if failed_images:
                    with st.expander(f"❌ Детали неудачных загрузок ({len(failed_images)})", expanded=False):
                        for error in failed_images:
                            st.text(f"• {error}")
        
        return True, ""
        
    except Exception as e:
        return False, f"Ошибка при обновлении файла: {str(e)}"
    
    finally:
        # Очищаем временные файлы изображений после сохранения
        if temp_image_files:
            cleaned_count = 0
            cleanup_errors = []
            
            for temp_file_path in temp_image_files:
                try:
                    if os.path.exists(temp_file_path):
                        os.unlink(temp_file_path)
                        cleaned_count += 1
                    else:
                        cleanup_errors.append(f"Файл не найден: {os.path.basename(temp_file_path)}")
                except Exception as cleanup_error:
                    cleanup_errors.append(f"Ошибка удаления {os.path.basename(temp_file_path)}: {cleanup_error}")
            
            # Компактный отчет об очистке
            if cleaned_count == len(temp_image_files):
                st.info(f"🧹 Очищено {cleaned_count} временных файлов")
            else:
                st.warning(f"🧹 Очищено {cleaned_count} из {len(temp_image_files)} временных файлов")
                
                # Показываем ошибки очистки в спойлере
                if cleanup_errors:
                    with st.expander(f"⚠️ Проблемы при очистке ({len(cleanup_errors)})", expanded=False):
                        for error in cleanup_errors:
                            st.text(f"• {error}")
        # Если нет временных файлов, не показываем ничего (убираем лишний лог)

def process_analytic_report(db_conn, file_path: str, include_images: bool = False) -> Tuple[bool, str, Dict]:
    """
    Main function to process the entire analytic report.
    
    Args:
        db_conn: Database connection
        file_path: Path to the Excel file
        include_images: Whether to download and insert product images from WB
        
    Returns:
        Tuple of (success, error_message, statistics_dict)
    """
    # Load and validate file
    df, wb, error_msg = load_analytic_report_file(file_path)
    if df is None:
        return False, error_msg, {}
    
    wb_sku_list = [str(sku) for sku in df['WB_SKU'].dropna().unique()]
    
    if not wb_sku_list:
        return False, "Нет валидных WB SKU для обработки", {}
    
    # Process each WB SKU
    wb_sku_data = {}
    
    # Get size mappings for all WB SKUs at once
    size_mappings = map_wb_to_ozon_by_size(db_conn, wb_sku_list)
    
    # Get Punta data for all WB SKUs at once
    punta_mappings = get_punta_data(db_conn, wb_sku_list)
    
    for wb_sku in wb_sku_list:
        size_mapping = size_mappings.get(wb_sku, {})
        
        # Get all oz_skus for this wb_sku
        all_oz_skus = []
        for oz_skus_list in size_mapping.values():
            all_oz_skus.extend(oz_skus_list)
        
        # Calculate size range
        size_range = calculate_size_range(size_mapping)
        
        # Get stock data
        total_stock = aggregate_stock_data(db_conn, all_oz_skus)
        
        # Get order statistics
        orders_data = generate_order_statistics(db_conn, all_oz_skus, days_back=30)
        
        # Get Punta data
        punta_data = punta_mappings.get(wb_sku, {})
        
        wb_sku_data[wb_sku] = {
            'sizes': size_mapping,
            'size_range': size_range,
            'total_stock': total_stock,
            'orders': orders_data,
            'punta': punta_data
        }
    
    # Update the file
    success, error_msg = update_analytic_report(file_path, wb_sku_data, include_images)
    
    if success:
        stats = {
            'processed_wb_skus': len(wb_sku_list),
            'wb_skus_with_ozon_matches': len([wb for wb, data in wb_sku_data.items() if data['sizes']]),
            'total_ozon_skus_found': sum(len([sku for skus_list in data['sizes'].values() for sku in skus_list]) for data in wb_sku_data.values()),
            'total_stock': sum(data['total_stock'] for data in wb_sku_data.values()),
            'wb_skus_with_punta_data': len([wb for wb, data in wb_sku_data.items() if any(data['punta'].values())])
        }
        return True, "", stats
    else:
        return False, error_msg, {}

def get_punta_data(db_conn, wb_sku_list: List[str]) -> Dict[str, Dict[str, str]]:
    """
    Retrieves Punta data for given WB SKUs.
    Now works with dynamic punta_table schema - automatically detects available columns.
    Takes FIRST occurrence for each wb_sku (most actual data) and formats values properly for Excel.
    
    Args:
        db_conn: Database connection
        wb_sku_list: List of WB SKUs to fetch data for
        
    Returns:
        Dict mapping wb_sku -> {column_name: value}
    """
    if not wb_sku_list:
        return {}
    
    try:
        # Import the new function to get dynamic columns
        from .db_crud import get_punta_table_columns
        
        # Get all available columns dynamically
        available_columns = get_punta_table_columns(db_conn)
        
        if not available_columns:
            st.info("📋 Таблица punta_table не существует или пуста")
            return {}
        
        # Remove wb_sku from list of data columns (it's the key column)
        data_columns = [col for col in available_columns if col != 'wb_sku']
        
        if not data_columns:
            st.warning("⚠️ В таблице punta_table нет колонок данных (только wb_sku)")
            return {}
        
        st.info(f"🔍 Найдено {len(data_columns)} колонок в punta_table: {', '.join(data_columns)}")
        
        # Convert wb_sku_list to integers for proper matching with punta_table
        wb_skus_int = []
        for wb_sku in wb_sku_list:
            try:
                wb_skus_int.append(int(wb_sku))
            except (ValueError, TypeError):
                # Skip invalid wb_sku values
                st.warning(f"Невалидный WB SKU для поиска в Punta: {wb_sku}")
                continue
        
        if not wb_skus_int:
            return {}
        
        # Build query to get punta data with dynamic columns
        # Use window function to get FIRST occurrence for each wb_sku
        if 'wb_sku' in available_columns:
            select_columns = ['wb_sku'] + data_columns
            
            punta_query = f"""
            WITH first_occurrences AS (
                SELECT {', '.join(f'"{col}"' for col in select_columns)},
                       ROW_NUMBER() OVER (PARTITION BY wb_sku ORDER BY ROWID) as rn
                FROM punta_table 
                WHERE wb_sku IN ({', '.join(['?'] * len(wb_skus_int))})
            )
            SELECT {', '.join(f'"{col}"' for col in select_columns)}
            FROM first_occurrences 
            WHERE rn = 1
            """
            
            punta_df = db_conn.execute(punta_query, wb_skus_int).fetchdf()
            
            if punta_df.empty:
                st.info(f"📭 Данные Punta не найдены для {len(wb_skus_int)} WB SKU")
                return {}
            
            st.success(f"✅ Найдены данные Punta для {len(punta_df)} из {len(wb_skus_int)} WB SKU (первые вхождения)")
            
            # Convert to dict format with Excel-friendly formatting
            result_map = {}
            for _, row in punta_df.iterrows():
                wb_sku = str(row['wb_sku'])  # Convert back to string for consistency with input
                result_map[wb_sku] = {}
                
                for column in data_columns:
                    if column in row and pd.notna(row[column]):
                        raw_value = row[column]
                        
                        # Format values to be Excel-friendly
                        if column == 'sort':
                            # Convert sort to integer string to avoid decimals in Excel
                            try:
                                if pd.notna(raw_value) and raw_value != '' and str(raw_value).lower() != 'nan':
                                    formatted_value = str(int(float(raw_value)))
                                else:
                                    formatted_value = ""
                            except (ValueError, TypeError):
                                formatted_value = str(raw_value) if raw_value else ""
                        elif column == 'supply_n':
                            # Handle supply_n: keep only numeric values, exclude dates
                            try:
                                str_value = str(raw_value).strip()
                                # If it looks like a date (contains dots or slashes), make it empty
                                if '.' in str_value or '/' in str_value or len(str_value) > 4:
                                    formatted_value = ""
                                else:
                                    # Try to parse as integer to ensure it's numeric
                                    int(str_value)
                                    formatted_value = str_value
                            except (ValueError, TypeError):
                                formatted_value = ""
                        else:
                            # For other columns, convert to string as usual
                            formatted_value = str(raw_value).strip()
                            # Clean up common problematic values
                            if formatted_value.lower() in ['nan', 'none', 'null']:
                                formatted_value = ""
                        
                        result_map[wb_sku][column] = formatted_value
                    else:
                        result_map[wb_sku][column] = ""
            
            return result_map
        else:
            st.error("❌ Колонка wb_sku не найдена в таблице punta_table")
            return {}
        
    except Exception as e:
        st.warning(f"Ошибка при получении данных Punta: {e}")
        return {}

def get_dynamic_punta_columns(db_conn) -> List[str]:
    """
    Получает список доступных колонок из punta_table для создания PUNTA_ колонок.
    
    Returns:
        List of column names available in punta_table (excluding wb_sku)
    """
    try:
        from .db_crud import get_punta_table_columns
        
        all_columns = get_punta_table_columns(db_conn)
        
        # Remove wb_sku as it's the key column, not data column
        data_columns = [col for col in all_columns if col != 'wb_sku']
        
        return data_columns
        
    except Exception as e:
        st.warning(f"Не удалось получить динамические колонки Punta: {e}")
        return []

def update_punta_columns_dynamically(ws, column_map: Dict[str, int], wb_sku_data: Dict[str, Dict], 
                                   data_start_row: int = 9):
    """
    Динамически обновляет все PUNTA_ колонки в Excel файле.
    Автоматически обрабатывает любые колонки, которые начинаются с PUNTA_.
    
    Args:
        ws: Excel worksheet object
        column_map: Mapping of column names to column numbers
        wb_sku_data: Data for each WB SKU
        data_start_row: Row number where data starts (default 9)
    """
    # Find all PUNTA_ columns in the Excel file
    punta_columns = {col_name: col_num for col_name, col_num in column_map.items() 
                     if col_name.startswith('PUNTA_')}
    
    if not punta_columns:
        return  # Тихо выходим если нет PUNTA_ колонок
    
    # Process each row
    for row_num in range(data_start_row, ws.max_row + 1):
        wb_sku_cell = ws.cell(row=row_num, column=column_map.get("WB_SKU", 2))
        wb_sku = str(wb_sku_cell.value).strip() if wb_sku_cell.value else ""
        
        if not wb_sku or wb_sku not in wb_sku_data:
            continue
        
        punta_data = wb_sku_data[wb_sku].get('punta', {})
        
        # Update each PUNTA_ column
        for col_name, col_num in punta_columns.items():
            # Extract the actual column name (remove PUNTA_ prefix)
            punta_col_name = col_name[6:]  # Remove 'PUNTA_' prefix
            punta_value = punta_data.get(punta_col_name, '')
            ws.cell(row=row_num, column=col_num).value = punta_value 