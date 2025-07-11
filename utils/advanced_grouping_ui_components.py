"""UI компоненты для улучшенной страницы группировки товаров.

Этот модуль содержит переиспользуемые UI компоненты для страницы
улучшенной группировки товаров, следуя паттернам проекта.

Автор: DataFox SL Project
Версия: 2.0.0
"""

import streamlit as st
import pandas as pd
from typing import List, Dict, Any, Optional
from utils.advanced_product_grouper import GroupingConfig, GroupingResult
from utils.wb_photo_service import get_wb_photo_url


def get_table_css() -> str:
    """Возвращает CSS стили для таблиц с фотографиями."""
    return """
    <style>
    .product-table {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        border-collapse: collapse;
        width: 100%;
        margin: 10px 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    
    .product-table th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: 600;
        padding: 12px 8px;
        text-align: left;
        font-size: 14px;
    }
    
    .product-table td {
        padding: 10px 8px;
        border-bottom: 1px solid #e9ecef;
        vertical-align: middle;
        font-size: 13px;
    }
    
    .product-table tr:hover {
        background-color: #f8f9fa;
    }
    
    .product-table tr:nth-child(even) {
        background-color: #fbfcfd;
    }
    
    .product-table td:first-child {
        text-align: center;
        width: 80px;
    }
    
    .product-table img {
        transition: transform 0.2s ease;
        cursor: pointer;
    }
    
    .product-table img:hover {
        transform: scale(1.1);
    }
    </style>
    """


@st.cache_data(ttl=3600)  # Кэшируем на 1 час
def get_photo_urls_batch(wb_skus: List[str]) -> List[str]:
    """Оптимизированное получение URL фотографий для списка WB SKU.
    
    Args:
        wb_skus: Список WB SKU
        
    Returns:
        List[str]: Список HTML элементов с фотографиями
    """
    photo_urls = []
    for sku in wb_skus:
        try:
            photo_url = get_wb_photo_url(sku)
            if photo_url:
                # Создаем HTML для изображения с lazy loading
                img_html = (f'<img src="{photo_url}" width="60" height="60" '
                          f'style="object-fit: cover; border-radius: 5px; '
                          f'border: 1px solid #ddd;" loading="lazy" '
                          f'alt="Товар {sku}" title="WB SKU: {sku}">')
                photo_urls.append(img_html)
            else:
                photo_urls.append('<div style="width:60px;height:60px;background:#f0f0f0;border-radius:5px;display:flex;align-items:center;justify-content:center;font-size:10px;">🚫 Нет фото</div>')
        except Exception as e:
            photo_urls.append('<div style="width:60px;height:60px;background:#ffe6e6;border-radius:5px;display:flex;align-items:center;justify-content:center;font-size:10px;">❌ Ошибка</div>')
    
    return photo_urls


def render_styled_table_with_photos(df: pd.DataFrame, table_id: str = "product-table"):
    """Отображает стилизованную таблицу с фотографиями.
    
    Args:
        df: DataFrame для отображения
        table_id: ID для CSS класса таблицы
    """
    if df.empty:
        st.warning("Нет данных для отображения")
        return
    
    # Простой HTML без сложных стилей
    try:
        # Генерируем упрощенный HTML
        html_parts = ["<div style='margin: 10px 0;'>"]
        html_parts.append("<table style='width: 100%; border-collapse: collapse; font-family: sans-serif; font-size: 13px;'>")
        
        # Заголовки
        html_parts.append("<thead style='background-color: #f0f2f6;'>")
        html_parts.append("<tr>")
        for col in df.columns:
            html_parts.append(f"<th style='padding: 8px; border: 1px solid #ddd; text-align: left;'>{col}</th>")
        html_parts.append("</tr>")
        html_parts.append("</thead>")
        
        # Данные
        html_parts.append("<tbody>")
        for _, row in df.iterrows():
            html_parts.append("<tr style='border-bottom: 1px solid #eee;'>")
            for col in df.columns:
                value = row[col]
                if col == "Фото":
                    # Специальная обработка для колонки с фотографиями
                    html_parts.append(f"<td style='padding: 4px; text-align: center; border: 1px solid #ddd;'>{value}</td>")
                else:
                    html_parts.append(f"<td style='padding: 8px; border: 1px solid #ddd;'>{value}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody>")
        
        html_parts.append("</table>")
        html_parts.append("</div>")
        
        # Объединяем и отображаем
        full_html = "".join(html_parts)
        st.html(full_html)
        
    except Exception as e:
        st.error(f"Ошибка отображения таблицы: {str(e)}")
        # Fallback - используем обычный st.dataframe без фотографий
        st.dataframe(df, use_container_width=True)


def render_grouping_configuration() -> GroupingConfig:
    """Отображает UI для настройки параметров группировки.
    
    Returns:
        GroupingConfig: Конфигурация группировки
    """
    st.subheader("⚙️ Настройки группировки")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Основные параметры:**")
        
        # Колонки для группировки
        available_columns = ['gender', 'wb_category', 'brand', 'color']
        grouping_columns = st.multiselect(
            "Колонки для группировки",
            options=available_columns,
            default=['gender', 'wb_category'],
            help="Товары будут группироваться по выбранным параметрам"
        )
        
        # Минимальный рейтинг группы
        min_group_rating = st.number_input(
            "Минимальный рейтинг группы",
            min_value=0.0,
            max_value=5.0,
            value=4.0,
            step=0.1,
            help="Группы с рейтингом ниже этого значения будут дополнены компенсаторами"
        )
        
        # Максимальное количество SKU в группе
        max_wb_sku_per_group = st.number_input(
            "Максимальное количество SKU в группе",
            min_value=1,
            max_value=50,
            value=10,
            help="Максимальное количество товаров в одной группе"
        )
    
    with col2:
        st.markdown("**Дополнительные опции:**")
        
        # Включить приоритет по полю sort
        enable_sort_priority = st.checkbox(
            "Использовать приоритет по полю 'sort'",
            value=True,
            help="Товары будут обрабатываться в порядке приоритета из поля sort"
        )
        
        # Фильтр по категории
        wb_category = st.text_input(
            "Фильтр по категории WB (опционально)",
            value="",
            help="Оставьте пустым для обработки всех категорий"
        )
        
        if wb_category.strip() == "":
            wb_category = None
    
    # Валидация
    if not grouping_columns:
        st.error("❌ Необходимо выбрать хотя бы одну колонку для группировки")
        st.stop()
    
    return GroupingConfig(
        grouping_columns=grouping_columns,
        min_group_rating=min_group_rating,
        max_wb_sku_per_group=max_wb_sku_per_group,
        enable_sort_priority=enable_sort_priority,
        wb_category=wb_category
    )


def render_wb_sku_input() -> List[str]:
    """Отображает UI для ввода WB SKU.
    
    Returns:
        List[str]: Список WB SKU для обработки
    """
    st.subheader("📝 Ввод WB SKU")
    
    input_method = st.radio(
        "Способ ввода WB SKU:",
        options=["Текстовое поле", "Загрузка файла"],
        horizontal=True
    )
    
    wb_skus = []
    
    if input_method == "Текстовое поле":
        wb_sku_text = st.text_area(
            "Введите WB SKU (по одному на строку):",
            height=150,
            help="Каждый WB SKU должен быть на отдельной строке"
        )
        
        if wb_sku_text.strip():
            wb_skus = [sku.strip() for sku in wb_sku_text.strip().split('\n') if sku.strip()]
    
    else:  # Загрузка файла
        uploaded_file = st.file_uploader(
            "Загрузите файл с WB SKU",
            type=['txt', 'csv', 'xlsx'],
            help="Файл должен содержать WB SKU в первой колонке"
        )
        
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.txt'):
                    content = uploaded_file.read().decode('utf-8')
                    wb_skus = [sku.strip() for sku in content.split('\n') if sku.strip()]
                
                elif uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                    wb_skus = df.iloc[:, 0].astype(str).tolist()
                
                elif uploaded_file.name.endswith('.xlsx'):
                    df = pd.read_excel(uploaded_file)
                    wb_skus = df.iloc[:, 0].astype(str).tolist()
                
                # Очищаем от пустых значений
                wb_skus = [sku.strip() for sku in wb_skus if str(sku).strip() and str(sku).strip() != 'nan']
                
            except Exception as e:
                st.error(f"❌ Ошибка при чтении файла: {str(e)}")
                wb_skus = []
    
    # Показываем количество найденных SKU
    if wb_skus:
        st.success(f"✅ Найдено {len(wb_skus)} WB SKU для обработки")
        
        # Показываем первые несколько для проверки
        with st.expander("Предварительный просмотр WB SKU"):
            preview_count = min(10, len(wb_skus))
            st.write(f"Первые {preview_count} из {len(wb_skus)} WB SKU:")
            for i, sku in enumerate(wb_skus[:preview_count], 1):
                st.write(f"{i}. {sku}")
            
            if len(wb_skus) > preview_count:
                st.write(f"... и еще {len(wb_skus) - preview_count} SKU")
    
    return wb_skus


def render_grouping_results(result: GroupingResult):
    """Отображает результаты группировки.
    
    Args:
        result: Результат группировки товаров
    """
    st.subheader("📊 Результаты группировки")
    
    # Статистика
    render_grouping_statistics(result.statistics)
    
    # Логи процесса
    render_process_logs(result.logs)
    
    # Группы товаров
    render_product_groups(result.groups)
    
    # Проблемные товары
    render_problematic_items(result.low_rating_items, result.defective_items)


def render_grouping_statistics(statistics: Dict[str, Any]):
    """Отображает статистику группировки.
    
    Args:
        statistics: Словарь со статистикой
    """
    st.markdown("### 📈 Статистика")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Всего групп",
            statistics.get('total_groups', 0)
        )
    
    with col2:
        st.metric(
            "Обработано товаров",
            statistics.get('total_items_processed', 0)
        )
    
    with col3:
        avg_size = statistics.get('avg_group_size', 0)
        st.metric(
            "Средний размер группы",
            f"{avg_size:.1f}"
        )
    
    with col4:
        avg_rating = statistics.get('avg_group_rating', 0)
        st.metric(
            "Средний рейтинг групп",
            f"{avg_rating:.2f}"
        )
    
    # Дополнительная статистика
    if statistics.get('low_rating_items_count', 0) > 0 or statistics.get('defective_items_count', 0) > 0:
        st.markdown("**Проблемные товары:**")
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric(
                "Товары с низким рейтингом",
                statistics.get('low_rating_items_count', 0)
            )
        
        with col2:
            st.metric(
                "Дефектные товары",
                statistics.get('defective_items_count', 0)
            )


def render_process_logs(logs: List[str]):
    """Отображает логи процесса группировки.
    
    Args:
        logs: Список сообщений лога
    """
    with st.expander("📋 Логи процесса группировки", expanded=False):
        if logs:
            for log_entry in logs:
                if "[ERROR]" in log_entry:
                    st.error(log_entry)
                elif "[WARNING]" in log_entry:
                    st.warning(log_entry)
                else:
                    st.info(log_entry)
        else:
            st.info("Логи отсутствуют")


def render_product_groups(groups: List[Dict[str, Any]]):
    """Отображает созданные группы товаров.
    
    Args:
        groups: Список групп товаров
    """
    st.markdown("### 🎯 Созданные группы")
    
    if not groups:
        st.warning("Группы не созданы")
        return
    
    # Фильтры для групп
    col1, col2 = st.columns(2)
    
    with col1:
        min_rating_filter = st.number_input(
            "Минимальный рейтинг для отображения",
            min_value=0.0,
            max_value=5.0,
            value=0.0,
            step=0.1,
            key="group_rating_filter"
        )
    
    with col2:
        min_size_filter = st.number_input(
            "Минимальный размер группы",
            min_value=1,
            max_value=50,
            value=1,
            key="group_size_filter"
        )
    
    # Фильтруем группы
    filtered_groups = [
        group for group in groups
        if group['group_rating'] >= min_rating_filter and
           group['item_count'] >= min_size_filter
    ]
    
    st.write(f"Показано {len(filtered_groups)} из {len(groups)} групп")
    
    # Отображаем группы
    for i, group in enumerate(filtered_groups, 1):
        with st.expander(
            f"Группа {group['group_id']}: {group['item_count']} товаров, рейтинг {group['group_rating']:.2f}",
            expanded=False
        ):
            # Отображаем информацию о группе
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Рейтинг группы", f"{group['group_rating']:.2f}")
            with col2:
                st.metric("Количество товаров", group['item_count'])
            with col3:
                priority_count = sum(1 for item in group['items'] if item.get('is_priority_item', False))
                st.metric("Приоритетных товаров", priority_count)
            
            # Создаем DataFrame для отображения
            items_df = pd.DataFrame(group['items'])
            
            # Выбираем колонки для отображения
            display_columns = ['wb_sku', 'avg_rating', 'total_stock', 'is_priority_item']
            if 'gender' in items_df.columns:
                display_columns.append('gender')
            if 'wb_category' in items_df.columns:
                display_columns.append('wb_category')
            if 'brand' in items_df.columns:
                display_columns.append('brand')
            if 'color' in items_df.columns:
                display_columns.append('color')
            
            # Фильтруем существующие колонки
            available_columns = [col for col in display_columns if col in items_df.columns]
            
            if available_columns:
                # Подготавливаем данные для отображения
                display_df = items_df[available_columns].copy()
                
                # Добавляем столбец с URL изображений в начало
                if 'wb_sku' in display_df.columns:
                    wb_skus = display_df['wb_sku'].astype(str).tolist()
                    # Используем оптимизированную функцию пакетного получения фото
                    photo_urls = get_photo_urls_batch(wb_skus)
                    # Вставляем столбец с фото в начало
                    display_df.insert(0, 'Фото', photo_urls)
                
                # Сохраняем оригинальную колонку для стилизации
                original_priority = display_df['is_priority_item'].copy() if 'is_priority_item' in display_df.columns else None
                
                # Переименовываем колонки для лучшего отображения
                column_names = {
                    'wb_sku': 'WB SKU',
                    'avg_rating': 'Рейтинг',
                    'total_stock': 'Остаток',
                    'is_priority_item': 'Приоритетный',
                    'gender': 'Пол',
                    'wb_category': 'Категория',
                    'brand': 'Бренд',
                    'color': 'Цвет'
                }
                display_df = display_df.rename(columns=column_names)
                
                # Форматируем колонку приоритетности
                if 'Приоритетный' in display_df.columns:
                    display_df['Приоритетный'] = display_df['Приоритетный'].astype(bool).map({True: '✅ Да', False: '❌ Нет'})
                
                # Создаем функцию стилизации с использованием оригинальных данных
                def highlight_priority_items(row):
                    if original_priority is not None:
                        row_index = row.name
                        # Проверяем, есть ли этот индекс в оригинальных данных
                        if row_index in original_priority.index and original_priority.loc[row_index]:
                            return ['background-color: #ffcdd2; color: #d32f2f; font-weight: bold'] * len(row)
                    return [''] * len(row)
                
                # Применяем стилизацию
                styled_df = display_df.style.apply(highlight_priority_items, axis=1)
                
                # Отображаем стилизованную таблицу с фотографиями
                render_styled_table_with_photos(display_df, f"group-{group['group_id']}")
                
                # Дополнительная информация о группе
                st.markdown("**Легенда:**")
                st.markdown("🔴 **Красный фон** - приоритетные товары (сезон 13 или положительный остаток)")
                st.markdown("⚪ **Обычный фон** - компенсаторы (неприоритетные товары с высоким рейтингом)")
            else:
                st.write("Данные для отображения недоступны")


def render_problematic_items(
    low_rating_items: List[Dict[str, Any]], 
    defective_items: List[Dict[str, Any]]
):
    """Отображает проблемные товары.
    
    Args:
        low_rating_items: Товары с низким рейтингом
        defective_items: Дефектные товары
    """
    if not low_rating_items and not defective_items:
        return
    
    st.markdown("### ⚠️ Проблемные товары")
    
    if low_rating_items:
        with st.expander(f"Товары с низким рейтингом ({len(low_rating_items)})", expanded=False):
            df = pd.DataFrame(low_rating_items)
            display_columns = ['wb_sku', 'avg_rating', 'is_priority_item']
            if 'gender' in df.columns:
                display_columns.append('gender')
            if 'wb_category' in df.columns:
                display_columns.append('wb_category')
            if 'total_stock' in df.columns:
                display_columns.append('total_stock')
            
            available_columns = [col for col in display_columns if col in df.columns]
            if available_columns:
                # Подготавливаем данные для отображения
                display_df = df[available_columns].copy()
                
                # Добавляем столбец с URL изображений в начало
                if 'wb_sku' in display_df.columns:
                    wb_skus = display_df['wb_sku'].astype(str).tolist()
                    # Используем оптимизированную функцию пакетного получения фото (меньший размер)
                    photo_urls = [url.replace('width="60" height="60"', 'width="50" height="50"') 
                                for url in get_photo_urls_batch(wb_skus)]
                    # Вставляем столбец с фото в начало
                    display_df.insert(0, 'Фото', photo_urls)
                
                # Переименовываем колонки
                column_names = {
                    'wb_sku': 'WB SKU',
                    'avg_rating': 'Рейтинг',
                    'total_stock': 'Остаток',
                    'is_priority_item': 'Приоритетный',
                    'gender': 'Пол',
                    'wb_category': 'Категория'
                }
                display_df = display_df.rename(columns=column_names)
                
                # Форматируем колонку приоритетности
                if 'Приоритетный' in display_df.columns:
                    display_df['Приоритетный'] = display_df['Приоритетный'].map({True: '✅ Да', False: '❌ Нет'})
                
                # Отображаем стилизованную таблицу с фотографиями
                render_styled_table_with_photos(display_df, "low-rating-items")
                
                # Показываем статистику
                priority_count = sum(1 for item in low_rating_items if item.get('is_priority_item', False))
                st.info(f"Из них приоритетных товаров: {priority_count}")
    
    if defective_items:
        with st.expander(f"Дефектные товары ({len(defective_items)})", expanded=False):
            df = pd.DataFrame(defective_items)
            display_columns = ['wb_sku', 'avg_rating', 'is_priority_item']
            if 'gender' in df.columns:
                display_columns.append('gender')
            if 'wb_category' in df.columns:
                display_columns.append('wb_category')
            if 'total_stock' in df.columns:
                display_columns.append('total_stock')
            
            available_columns = [col for col in display_columns if col in df.columns]
            if available_columns:
                # Подготавливаем данные для отображения
                display_df = df[available_columns].copy()
                
                # Добавляем столбец с URL изображений в начало
                if 'wb_sku' in display_df.columns:
                    wb_skus = display_df['wb_sku'].astype(str).tolist()
                    # Используем оптимизированную функцию пакетного получения фото (меньший размер)
                    photo_urls = [url.replace('width="60" height="60"', 'width="50" height="50"') 
                                for url in get_photo_urls_batch(wb_skus)]
                    # Вставляем столбец с фото в начало
                    display_df.insert(0, 'Фото', photo_urls)
                
                # Переименовываем колонки
                column_names = {
                    'wb_sku': 'WB SKU',
                    'avg_rating': 'Рейтинг',
                    'total_stock': 'Остаток',
                    'is_priority_item': 'Приоритетный',
                    'gender': 'Пол',
                    'wb_category': 'Категория'
                }
                display_df = display_df.rename(columns=column_names)
                
                # Форматируем колонку приоритетности
                if 'Приоритетный' in display_df.columns:
                    display_df['Приоритетный'] = display_df['Приоритетный'].map({True: '✅ Да', False: '❌ Нет'})
                
                # Отображаем стилизованную таблицу с фотографиями
                render_styled_table_with_photos(display_df, "defective-items")
                
                # Показываем статистику
                priority_count = sum(1 for item in defective_items if item.get('is_priority_item', False))
                st.info(f"Из них приоритетных товаров: {priority_count}")


def render_export_options(result: GroupingResult):
    """Отображает опции экспорта результатов.
    
    Args:
        result: Результат группировки
    """
    st.subheader("💾 Экспорт результатов")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Экспорт в Excel", use_container_width=True):
            excel_data = export_to_excel(result)
            if excel_data:
                st.download_button(
                    label="⬇️ Скачать Excel файл",
                    data=excel_data,
                    file_name=f"product_groups_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    
    with col2:
        if st.button("📋 Экспорт в CSV", use_container_width=True):
            csv_data = export_to_csv(result)
            if csv_data:
                st.download_button(
                    label="⬇️ Скачать CSV файл",
                    data=csv_data,
                    file_name=f"product_groups_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )


def export_to_excel(result: GroupingResult) -> Optional[bytes]:
    """Экспортирует результаты в Excel формат.
    
    Args:
        result: Результат группировки
        
    Returns:
        bytes: Данные Excel файла или None при ошибке
    """
    try:
        import io
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Лист со статистикой
        ws_stats = wb.active
        ws_stats.title = "Статистика"
        
        stats_data = [
            ["Параметр", "Значение"],
            ["Всего групп", result.statistics.get('total_groups', 0)],
            ["Обработано товаров", result.statistics.get('total_items_processed', 0)],
            ["Средний размер группы", f"{result.statistics.get('avg_group_size', 0):.1f}"],
            ["Средний рейтинг групп", f"{result.statistics.get('avg_group_rating', 0):.2f}"],
            ["Товары с низким рейтингом", result.statistics.get('low_rating_items_count', 0)],
            ["Дефектные товары", result.statistics.get('defective_items_count', 0)]
        ]
        
        for row in stats_data:
            ws_stats.append(row)
        
        # Лист с группами
        ws_groups = wb.create_sheet("Группы")
        
        # Заголовки
        headers = ["Группа", "WB SKU", "Рейтинг", "Остатки", "Пол", "Категория"]
        ws_groups.append(headers)
        
        # Данные групп
        for group in result.groups:
            for item in group['items']:
                row = [
                    group['group_id'],
                    item.get('wb_sku', ''),
                    item.get('avg_rating', ''),
                    item.get('total_stock', ''),
                    item.get('gender', ''),
                    item.get('wb_category', '')
                ]
                ws_groups.append(row)
        
        # Сохраняем в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Ошибка при экспорте в Excel: {str(e)}")
        return None


def export_to_csv(result: GroupingResult) -> Optional[str]:
    """Экспортирует результаты в CSV формат.
    
    Args:
        result: Результат группировки
        
    Returns:
        str: Данные CSV файла или None при ошибке
    """
    try:
        rows = []
        
        # Заголовки
        headers = ["Группа", "WB SKU", "Рейтинг", "Остатки", "Пол", "Категория"]
        rows.append(headers)
        
        # Данные групп
        for group in result.groups:
            for item in group['items']:
                row = [
                    group['group_id'],
                    item.get('wb_sku', ''),
                    item.get('avg_rating', ''),
                    item.get('total_stock', ''),
                    item.get('gender', ''),
                    item.get('wb_category', '')
                ]
                rows.append(row)
        
        # Преобразуем в CSV
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Ошибка при экспорте в CSV: {str(e)}")
        return None


def render_help_section():
    """Отображает раздел помощи."""
    with st.expander("❓ Справка по использованию", expanded=False):
        st.markdown("""
        ### Как использовать улучшенную группировку товаров:
        
        1. **Настройка параметров:**
           - Выберите колонки для группировки (товары с одинаковыми значениями будут объединены)
           - Установите минимальный рейтинг группы (группы с низким рейтингом будут дополнены компенсаторами)
           - Укажите максимальное количество SKU в группе
        
        2. **Ввод данных:**
           - Введите WB SKU через текстовое поле или загрузите файл
           - Поддерживаются форматы: TXT, CSV, XLSX
        
        3. **Запуск группировки:**
           - Нажмите кнопку "Создать группы"
           - Дождитесь завершения обработки
        
        4. **Анализ результатов:**
           - Изучите статистику группировки
           - Просмотрите созданные группы
           - Проверьте проблемные товары
        
        5. **Экспорт:**
           - Экспортируйте результаты в Excel или CSV для дальнейшей работы
        
        ### Улучшения по сравнению с оригинальной версией:
        
        - ✅ Исправлены SQL запросы для поиска компенсаторов
        - ✅ Справедливое распределение компенсаторов между группами
        - ✅ Детальное логирование процесса
        - ✅ Улучшенная архитектура кода
        - ✅ Расширенные возможности экспорта
        """)


def create_marketplace_summary_table(result: GroupingResult, connection) -> pd.DataFrame:
    """Создает итоговую таблицу для загрузки на маркетплейс.
    
    Структура таблицы:
    - wb_sku: артикул WB
    - oz_vendor_code: артикул поставщика на Озоне
    - объединяющий_код: код для объединения товаров
    
    Логика объединяющего кода:
    - Если товар приоритетный - объединяющий код = wb_sku этого товара
    - Если товар не приоритетный - объединяющий код = wb_sku приоритетного товара из той же группы
    
    Args:
        result: Результат группировки
        connection: Соединение с базой данных
        
    Returns:
        DataFrame с итоговой таблицей
    """
    try:
        from utils.cross_marketplace_linker import CrossMarketplaceLinker
        
        # Собираем все wb_sku из всех групп + товары с браком
        all_wb_skus = []
        for group in result.groups:
            for item in group['items']:
                all_wb_skus.append(item['wb_sku'])
        
        # Добавляем товары с браком
        for defective_item in result.defective_items:
            all_wb_skus.append(defective_item['wb_sku'])
        
        if not all_wb_skus:
            return pd.DataFrame(columns=['wb_sku', 'oz_vendor_code', 'объединяющий_код'])
        
        # Получаем связи wb_sku -> oz_vendor_code
        linker = CrossMarketplaceLinker(connection)
        extended_links = linker.get_extended_links(all_wb_skus, include_product_details=False)
        
        # Создаем итоговую таблицу
        marketplace_rows = []
        
        for group in result.groups:
            # Определяем приоритетный товар группы (главный wb_sku)
            main_wb_sku = group.get('main_wb_sku')
            
            # Если main_wb_sku не определен, берем первый приоритетный товар
            if not main_wb_sku:
                priority_items = [item for item in group['items'] if item.get('is_priority_item', False)]
                if priority_items:
                    main_wb_sku = priority_items[0]['wb_sku']
                else:
                    # Если нет приоритетных товаров, берем первый товар группы
                    main_wb_sku = group['items'][0]['wb_sku']
            
            # Обрабатываем все товары группы
            for item in group['items']:
                wb_sku = item['wb_sku']
                is_priority = item.get('is_priority_item', False)
                
                # Определяем объединяющий код
                if is_priority:
                    # Приоритетный товар - объединяющий код = его wb_sku
                    unifying_code = wb_sku
                else:
                    # Не приоритетный товар - объединяющий код = wb_sku приоритетного товара
                    unifying_code = main_wb_sku
                
                # Получаем все oz_vendor_code для этого wb_sku
                wb_links = extended_links[extended_links['wb_sku'] == wb_sku]
                
                if not wb_links.empty:
                    # Создаем строку для каждого oz_vendor_code
                    for _, link in wb_links.iterrows():
                        oz_vendor_code = link.get('oz_vendor_code', '')
                        
                        marketplace_rows.append({
                            'wb_sku': wb_sku,
                            'oz_vendor_code': oz_vendor_code,
                            'объединяющий_код': unifying_code
                        })
                else:
                    # Если нет связей с Озоном, добавляем строку с пустым oz_vendor_code
                    marketplace_rows.append({
                        'wb_sku': wb_sku,
                        'oz_vendor_code': '',
                        'объединяющий_код': unifying_code
                    })
        
        # НОВОЕ: Обрабатываем товары с браком отдельно
        for defective_item in result.defective_items:
            wb_sku = defective_item['wb_sku']
            
            # Для товаров с браком используем уникальный код объединения
            unifying_code = f"БракSH_{wb_sku}"
            
            # Получаем все oz_vendor_code для этого wb_sku
            wb_links = extended_links[extended_links['wb_sku'] == wb_sku]
            
            if not wb_links.empty:
                # Создаем строку для каждого oz_vendor_code
                for _, link in wb_links.iterrows():
                    oz_vendor_code = link.get('oz_vendor_code', '')
                    
                    marketplace_rows.append({
                        'wb_sku': wb_sku,
                        'oz_vendor_code': oz_vendor_code,
                        'объединяющий_код': unifying_code
                    })
            else:
                # Если нет связей с Озоном, добавляем строку с пустым oz_vendor_code
                marketplace_rows.append({
                    'wb_sku': wb_sku,
                    'oz_vendor_code': '',
                    'объединяющий_код': unifying_code
                })
        
        # Создаем DataFrame
        if marketplace_rows:
            marketplace_df = pd.DataFrame(marketplace_rows)
            
            # Убираем дубликаты (может быть несколько одинаковых строк)
            marketplace_df = marketplace_df.drop_duplicates()
            
            # Сортируем по объединяющему коду, затем по wb_sku
            marketplace_df = marketplace_df.sort_values(['объединяющий_код', 'wb_sku'])
            
            return marketplace_df
        else:
            return pd.DataFrame(columns=['wb_sku', 'oz_vendor_code', 'объединяющий_код'])
        
    except Exception as e:
        st.error(f"Ошибка при создании итоговой таблицы: {str(e)}")
        return pd.DataFrame(columns=['wb_sku', 'oz_vendor_code', 'объединяющий_код'])


def export_marketplace_summary_to_excel(marketplace_df: pd.DataFrame) -> Optional[bytes]:
    """Экспортирует итоговую таблицу для маркетплейса в Excel.
    
    Args:
        marketplace_df: DataFrame с итоговой таблицей
        
    Returns:
        bytes: Данные Excel файла или None при ошибке
    """
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Итоговая таблица"
        
        # Заголовки
        headers = ['WB SKU', 'OZ Vendor Code', 'Объединяющий код']
        ws.append(headers)
        
        # Стилизация заголовков
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for _, row in marketplace_df.iterrows():
            ws.append([
                row['wb_sku'],
                row['oz_vendor_code'],
                row['объединяющий_код']
            ])
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Ошибка при экспорте итоговой таблицы в Excel: {str(e)}")
        return None


def export_marketplace_summary_to_csv(marketplace_df: pd.DataFrame) -> Optional[str]:
    """Экспортирует итоговую таблицу для маркетплейса в CSV.
    
    Args:
        marketplace_df: DataFrame с итоговой таблицей
        
    Returns:
        str: Данные CSV файла или None при ошибке
    """
    try:
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Заголовки
        writer.writerow(['WB SKU', 'OZ Vendor Code', 'Объединяющий код'])
        
        # Данные
        for _, row in marketplace_df.iterrows():
            writer.writerow([
                row['wb_sku'],
                row['oz_vendor_code'],
                row['объединяющий_код']
            ])
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Ошибка при экспорте итоговой таблицы в CSV: {str(e)}")
        return None


def render_marketplace_summary_table(result: GroupingResult, connection):
    """Отображает итоговую таблицу для загрузки на маркетплейс.
    
    Args:
        result: Результат группировки
        connection: Соединение с базой данных
    """
    st.subheader("📋 Итоговая таблица для загрузки на маркетплейс")
    
    # Создаем итоговую таблицу
    with st.spinner("Формирование итоговой таблицы..."):
        marketplace_df = create_marketplace_summary_table(result, connection)
    
    if marketplace_df.empty:
        st.warning("⚠️ Не удалось создать итоговую таблицу - нет данных для экспорта")
        return
    
    # Показываем статистику
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Всего записей", len(marketplace_df))
    
    with col2:
        unique_wb_skus = marketplace_df['wb_sku'].nunique()
        st.metric("Уникальных WB SKU", unique_wb_skus)
    
    with col3:
        unique_oz_codes = marketplace_df[marketplace_df['oz_vendor_code'] != '']['oz_vendor_code'].nunique()
        st.metric("Уникальных OZ кодов", unique_oz_codes)
    
    with col4:
        unique_unifying_codes = marketplace_df['объединяющий_код'].nunique()
        st.metric("Уникальных групп", unique_unifying_codes)
    
    # Показываем превью таблицы
    st.markdown("**Превью таблицы (первые 20 строк):**")
    st.dataframe(marketplace_df.head(20), use_container_width=True)
    
    # Показываем пример группировки
    with st.expander("📊 Пример группировки по объединяющему коду", expanded=False):
        if len(marketplace_df) > 0:
            # Берем первую группу для примера
            first_group = marketplace_df['объединяющий_код'].iloc[0]
            group_example = marketplace_df[marketplace_df['объединяющий_код'] == first_group]
            
            st.write(f"**Группа с объединяющим кодом: {first_group}**")
            st.dataframe(group_example, use_container_width=True)
            
            st.write(f"В этой группе **{len(group_example)} товаров** объединены одним кодом")
    
    # Кнопки экспорта
    st.divider()
    st.markdown("**💾 Экспорт итоговой таблицы:**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("📊 Экспорт в Excel", key="export_marketplace_excel", use_container_width=True):
            excel_data = export_marketplace_summary_to_excel(marketplace_df)
            if excel_data:
                st.download_button(
                    label="⬇️ Скачать Excel файл",
                    data=excel_data,
                    file_name=f"marketplace_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_marketplace_excel"
                )
    
    with col2:
        if st.button("📋 Экспорт в CSV", key="export_marketplace_csv", use_container_width=True):
            csv_data = export_marketplace_summary_to_csv(marketplace_df)
            if csv_data:
                st.download_button(
                    label="⬇️ Скачать CSV файл",
                    data=csv_data,
                    file_name=f"marketplace_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="download_marketplace_csv"
                )