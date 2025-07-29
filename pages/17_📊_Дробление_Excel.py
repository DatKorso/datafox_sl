"""
Streamlit page for splitting Excel files into chunks based on unique values in a specified column.

This page allows users to:
- Upload an Excel file
- Select a sheet and column for splitting
- Specify the number of unique values per chunk
- Download a ZIP archive with the split files

The splitting logic:
1. Reads the specified column and finds unique values
2. Groups rows by unique values (all rows with the same value stay together)
3. Creates chunks with the specified number of unique values
4. Generates separate Excel files for each chunk
5. Packages all files into a ZIP archive

Author: DataFox SL Project
Version: 1.0.0
"""

import streamlit as st
import pandas as pd
import os
from io import BytesIO
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import zipfile
from typing import List, Dict, Any, Tuple
import time

# Page configuration
st.set_page_config(
    page_title="Дробление Excel",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Дробление Excel файлов")
st.markdown("Разделение Excel файла на части по уникальным значениям в выбранном столбце")
st.markdown("---")

# Initialize session state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'available_sheets' not in st.session_state:
    st.session_state.available_sheets = []
if 'selected_sheet' not in st.session_state:
    st.session_state.selected_sheet = None
if 'df_preview' not in st.session_state:
    st.session_state.df_preview = None
if 'column_names' not in st.session_state:
    st.session_state.column_names = []
if 'processing_config' not in st.session_state:
    st.session_state.processing_config = {}

def get_excel_sheets(file_bytes: bytes) -> List[str]:
    """Get list of sheets from Excel file"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()
            
            wb = load_workbook(tmp_file.name, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            os.unlink(tmp_file.name)
            return sheets
    except Exception as e:
        st.error(f"Ошибка при чтении файла: {str(e)}")
        return []

def read_excel_sheet(file_bytes: bytes, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    """Read Excel sheet with proper header handling"""
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(file_bytes)
            tmp_file.flush()
            
            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=header_row)
            os.unlink(tmp_file.name)
            return df
    except Exception as e:
        st.error(f"Ошибка при чтении листа {sheet_name}: {str(e)}")
        return pd.DataFrame()

def get_unique_values_count(df: pd.DataFrame, column_name: str) -> int:
    """Get count of unique values in a column"""
    try:
        if column_name not in df.columns:
            return 0
        return df[column_name].nunique()
    except Exception as e:
        st.error(f"Ошибка при подсчете уникальных значений: {str(e)}")
        return 0

def split_dataframe_by_unique_values(df: pd.DataFrame, column_name: str, chunk_size: int) -> List[pd.DataFrame]:
    """Split DataFrame into chunks based on unique values in specified column"""
    try:
        if column_name not in df.columns:
            st.error(f"Столбец '{column_name}' не найден в данных")
            return []
        
        # Get unique values
        unique_values = df[column_name].unique()
        unique_values = [val for val in unique_values if pd.notna(val)]  # Remove NaN values
        
        if not unique_values:
            st.error("Не найдено уникальных значений в выбранном столбце")
            return []
        
        chunks = []
        for i in range(0, len(unique_values), chunk_size):
            chunk_values = unique_values[i:i + chunk_size]
            chunk_df = df[df[column_name].isin(chunk_values)].copy()
            if not chunk_df.empty:
                chunks.append(chunk_df)
        
        return chunks
    except Exception as e:
        st.error(f"Ошибка при разделении данных: {str(e)}")
        return []

def create_zip_archive(chunks: List[pd.DataFrame], base_filename: str) -> BytesIO:
    """Create ZIP archive with Excel files for each chunk"""
    try:
        zip_buffer = BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i, chunk_df in enumerate(chunks, 1):
                # Create Excel file in memory
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    chunk_df.to_excel(writer, index=False, sheet_name='Data')
                
                # Add to ZIP
                filename = f"{base_filename}_chunk_{i}_of_{len(chunks)}.xlsx"
                zip_file.writestr(filename, excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        return zip_buffer
    except Exception as e:
        st.error(f"Ошибка при создании архива: {str(e)}")
        return None

# Main UI
st.subheader("1. Загрузка файла")
uploaded_file = st.file_uploader(
    "Выберите Excel файл для разделения",
    type=['xlsx', 'xls'],
    help="Поддерживаются файлы Excel (.xlsx, .xls)"
)

if uploaded_file is not None:
    st.session_state.uploaded_file = uploaded_file
    file_bytes = uploaded_file.read()
    
    # Get sheets
    sheets = get_excel_sheets(file_bytes)
    if sheets:
        st.session_state.available_sheets = sheets
        st.success(f"Файл загружен успешно. Найдено листов: {len(sheets)}")
        
        # Sheet selection
        st.subheader("2. Выбор листа")
        selected_sheet = st.selectbox(
            "Выберите лист для обработки",
            options=sheets,
            help="Выберите лист Excel для разделения"
        )
        
        if selected_sheet:
            st.session_state.selected_sheet = selected_sheet
            
            # Read and preview data
            with st.spinner("Загрузка данных..."):
                df = read_excel_sheet(file_bytes, selected_sheet)
                
            if not df.empty:
                st.session_state.df_preview = df
                st.session_state.column_names = list(df.columns)
                
                # Data preview
                st.subheader("3. Предварительный просмотр данных")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Общее количество строк", len(df))
                with col2:
                    st.metric("Количество столбцов", len(df.columns))
                with col3:
                    st.metric("Размер файла", f"{len(file_bytes) / 1024:.1f} KB")
                
                # Show preview
                st.dataframe(df.head(10), use_container_width=True)
                
                # Configuration
                st.subheader("4. Настройка разделения")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    selected_column = st.selectbox(
                        "Выберите столбец для группировки",
                        options=df.columns,
                        help="Выберите столбец, по уникальным значениям которого будет происходить разделение"
                    )
                
                with col2:
                    chunk_size = st.number_input(
                        "Количество уникальных значений на файл",
                        min_value=1,
                        max_value=1000,
                        value=100,
                        help="Количество уникальных значений, которые будут включены в каждый файл"
                    )
                
                if selected_column:
                    # Show statistics for selected column
                    unique_count = get_unique_values_count(df, selected_column)
                    estimated_files = (unique_count + chunk_size - 1) // chunk_size  # Ceiling division
                    
                    st.info(f"""
                    **Статистика по столбцу '{selected_column}':**
                    - Уникальных значений: {unique_count}
                    - Предполагаемое количество файлов: {estimated_files}
                    - Строк на файл: варьируется (зависит от частоты значений)
                    """)
                    
                    # Show sample unique values
                    with st.expander("Примеры уникальных значений"):
                        sample_values = df[selected_column].unique()[:20]
                        st.write(", ".join([str(v) for v in sample_values if pd.notna(v)]))
                        if len(df[selected_column].unique()) > 20:
                            st.write("... и еще", len(df[selected_column].unique()) - 20, "значений")
                    
                    # Processing button
                    st.subheader("5. Обработка")
                    
                    if st.button("🚀 Начать разделение файла", type="primary"):
                        if unique_count == 0:
                            st.error("В выбранном столбце нет уникальных значений для разделения")
                        else:
                            # Store configuration
                            st.session_state.processing_config = {
                                'column': selected_column,
                                'chunk_size': chunk_size,
                                'unique_count': unique_count,
                                'estimated_files': estimated_files
                            }
                            
                            # Process the file
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            try:
                                # Split data
                                status_text.text("Разделение данных...")
                                progress_bar.progress(0.3)
                                
                                chunks = split_dataframe_by_unique_values(df, selected_column, chunk_size)
                                
                                if chunks:
                                    status_text.text("Создание Excel файлов...")
                                    progress_bar.progress(0.6)
                                    
                                    # Create ZIP archive
                                    base_filename = os.path.splitext(uploaded_file.name)[0]
                                    zip_buffer = create_zip_archive(chunks, base_filename)
                                    
                                    if zip_buffer:
                                        progress_bar.progress(1.0)
                                        status_text.text("Готово!")
                                        
                                        # Success metrics
                                        st.success("Файл успешно разделен!")
                                        
                                        col1, col2, col3 = st.columns(3)
                                        with col1:
                                            st.metric("Создано файлов", len(chunks))
                                        with col2:
                                            st.metric("Обработано уникальных значений", unique_count)
                                        with col3:
                                            st.metric("Размер архива", f"{len(zip_buffer.getvalue()) / 1024:.1f} KB")
                                        
                                        # Download button
                                        st.download_button(
                                            label="📥 Скачать архив с разделенными файлами",
                                            data=zip_buffer.getvalue(),
                                            file_name=f"{base_filename}_split_files.zip",
                                            mime="application/zip",
                                            type="primary"
                                        )
                                        
                                        # Show chunk details
                                        with st.expander("Детали разделения"):
                                            for i, chunk in enumerate(chunks, 1):
                                                unique_in_chunk = chunk[selected_column].nunique()
                                                st.write(f"Файл {i}: {len(chunk)} строк, {unique_in_chunk} уникальных значений")
                                
                                else:
                                    st.error("Не удалось создать разделенные файлы")
                                    
                            except Exception as e:
                                st.error(f"Ошибка при обработке: {str(e)}")
                                with st.expander("Детали ошибки"):
                                    st.code(traceback.format_exc())
                            
                            finally:
                                progress_bar.empty()
                                status_text.empty()

else:
    st.info("👆 Загрузите Excel файл для начала работы")
    
    # Help section
    with st.expander("ℹ️ Справка"):
        st.markdown("""
        ### Как использовать:
        
        1. **Загрузите Excel файл** - поддерживаются форматы .xlsx и .xls
        2. **Выберите лист** для обработки
        3. **Выберите столбец** для группировки (по уникальным значениям которого будет происходить разделение)
        4. **Укажите количество уникальных значений** на файл
        5. **Нажмите кнопку обработки** и дождитесь создания архива
        6. **Скачайте ZIP архив** с разделенными файлами
        
        ### Принцип работы:
        - Все строки с одинаковыми значениями в выбранном столбце остаются в одном файле
        - Например, если значение "ABC123" встречается в 50 строках, все 50 строк попадут в один файл
        - Файлы создаются по принципу: N уникальных значений = 1 файл
        
        ### Пример:
        - У вас есть столбец "Артикул" с 500 уникальными значениями
        - Вы указываете 100 уникальных значений на файл
        - Получите 5 файлов, каждый с 100 уникальными артикулами (и всеми связанными строками)
        """)